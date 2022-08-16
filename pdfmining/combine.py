from openpyxl import Workbook, load_workbook
from fix import copyRow

dateInd = 21
dispInd = 19
firstInd = 3
middleInd = 4
lastInd = 5
countyInd = 0
codeInd = 23
severityInd = 29

def isDuplicate(newRow, sheet, log):
	duplicate = False
	for line, row in enumerate(list(sheet.rows), 1):
		for i, item in enumerate(list(row)):
			if newRow[i] != str(item.value): break
			if i+1 == len(list(row)): duplicate = True

		if duplicate: 
			log.write("ORIG: " + str(copyRow(row)) + "\n")
			log.write("DUPE: " + str(newRow) + "\n")
			log.write(str(line) + "\n\n\n")
			break

	return duplicate

def getData(original):
	dataRows = list()
	for row in list(original.rows):
		r = copyRow(row)

		date = r[dateInd]
		dispMethod = r[dispInd]
		firstName = r[firstInd]
		middleName = r[middleInd]
		lastName = r[lastInd]
		county = r[countyInd]
		code = r[codeInd]
		severity = r[severityInd]

		dataRows.append([date, dispMethod, firstName, middleName, lastName, county, code, severity])

	return dataRows

def main():
	mecBook= load_workbook(filename="fixed/MEC_ORIGINAL.xlsx", data_only=True)
	mecSheet = mecBook.active

	rows = getData(mecSheet)
	mecBook.close()

	compiled = load_workbook(filename="fixed/compiled.xlsx", data_only=True)
	compSheet = compiled.active

	log = open("fixed/log.txt", "w")
	for row in rows:
		if not isDuplicate(row, compSheet, log): compSheet.append(row)

	log.close()

	compiled.save(filename="fixed/final.xlsx")
	compiled.close()


if __name__ == "__main__": main()
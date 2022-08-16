from openpyxl import Workbook, load_workbook
import glob, re
from progressBar import progressbar

datePattern = re.compile("^(\d{1,2}/){1,2}\d{4}$")
docketPattern = re.compile("^\w{1,8}-\w{1,8}\W")
namePattern = re.compile("[a-zA-Z]+")
numberPattern = re.compile("^\d{1,3}$")
tablePattern = re.compile("^\w\W=\W")
districtPattern = re.compile("^District\W:\W\d+$")
countyColumn = "B"
nameColumn = "B"
cleanColumn = "B"
dateColumn = "J"
eventColumn = "H"
dispColumn = "I"
codeColumn = "L"
severityColumn = "O"

def getEndRow(sheet):
	return len(list(sheet.rows))

def getStartRow(sheet):
	startRow = 0
	for i, item in enumerate(sheet[dateColumn], 1):
		string = str(item.value)
		if string == "Sentence Date": return i+1

	print("ERROR: Bad startRow.")
	return 0

def getLines(sheet):	# regex get each line with a new date
	lines = list()
	for i, item in enumerate(sheet[dateColumn], 1): # iter all cells in date column
		string = str(item.value)
		match = re.search(datePattern, string)
		if match: lines.append(i)

	return lines

def getNumRows(sheet, lines, i):
	if i+1 == len(lines): return getEndRow(sheet)+1 - lines[i]

	for line in range(lines[i] + 1, lines[i+1]): # check for rogue county
		if re.search(numberPattern, str(sheet[countyColumn + str(line)].value)):
			return line - lines[i]
	
	return lines[i+1] - lines[i]

def copyRow(row):
	out = list()
	for item in row:
		if item.value is None: out.append(None)
		else: out.append(str(item.value))

	return out

def getDisposition(sheet, lines, i):
	numRows = getNumRows(sheet, lines, i)
	event = ""
	for item in list(sheet[eventColumn])[lines[i]-1:lines[i] + numRows-1]:
		if item.value is not None: event += str(item.value) + " "

	disposition = ""
	for item in list(sheet[dispColumn])[lines[i]-1:lines[i] + numRows-1]:
		if item.value is not None: disposition += str(item.value) + " "

	return event + "- " + disposition

def getCounty(sheet, lines, i):
	county = ""
	items = list(sheet[countyColumn])
	for r in range(lines[i], 0, -1): # used to be (-2, -1, -1)
		match = re.search(numberPattern, str(items[r-1].value))
		if match: county = str(items[r].value)

	return county

def cleanNameString(s): # removes bits of string that grind the software's gears
	string = s

	removeChars = [',', '-', '=']
	for c in removeChars: string = string.replace(c, '')

	spaceCheck = ['  ', '   ', '    ']
	for s in spaceCheck: string = string.replace(s, ' ')

	if string[0] == ' ': string = string[1:]
	if string[-1] == ' ': string = string[:-1]

	return string


def getNames(sheet, lines, i): # gets the full name for an item
	string = ""
	numRows = getNumRows(sheet, lines, i)

	for item in list(sheet[nameColumn])[lines[i]-1:lines[i] + numRows-1]:
		if item.value is not None: string += str(item.value) + " "
	string = string[:len(string)-1] # remove last space

	dMatch = re.search(docketPattern, string)
	if dMatch: string = string[len(dMatch.group(0)):]
	else: print("ERROR: getName() regex failure - " + "line: " + str(lines[i]))

	string = cleanNameString(string)

	splitString = string.split(" ", 2)

	nameList = [splitString[1], "", splitString[0]]
	if len(splitString) == 3: nameList[1] = splitString[2]

	return nameList

def removeExtraRows(sheet, i): # i is first bad row
	sheet.delete_rows(i, getEndRow(sheet) - i + 1)

def cleanFile(sheet):
	for i, item in reversed(list(enumerate(list(sheet[cleanColumn]), 1))): # reversed enumerated list of items to check
		string = str(item.value)
		dMatch = re.search(districtPattern, string)
		tMatch = re.search(tablePattern, string)

		if dMatch: 
			removeExtraRows(sheet, i)
			break
		if tMatch:
			removeExtraRows(sheet, i-1)
			break

def combineSheets(files, year):
	outbook = Workbook()
	outsheet = outbook.active

	for file in files:
		tmpbook = load_workbook(filename=file, data_only=True)
		tmpsheet = tmpbook.active
		
		cleanFile(tmpsheet)

		startRow = getStartRow(tmpsheet)
		rows = list(tmpsheet.rows)
		for r in rows[startRow-1:]:
			outsheet.append(copyRow(r))

		tmpbook.close()

	outbook.save(filename="xlsx_out/combined/" + str(year) + ".xlsx")
	outbook.close()

def parseFile(fileName, finalsheet):
	workbook = load_workbook(filename=str(fileName), data_only=True)
	sheet = workbook.active
	lines = getLines(sheet)
	
	for i, line in enumerate(lines): 	# get each item from xlsx file
		date = str(sheet[dateColumn + str(line)].value)
		dispMethod = getDisposition(sheet, lines, i)
		names = getNames(sheet, lines, i)
		firstName = names[0]
		middleName = names[1]
		lastName = names[2]
		county = getCounty(sheet, lines, i)
		code = str(sheet[codeColumn + str(line)].value)
		severity = str(sheet[severityColumn + str(line)].value)

		row = [date, dispMethod, firstName, middleName, lastName, county, code, severity] # compile data into list
		finalsheet.append(row)

	workbook.close()

def main():
	finalbook = Workbook()
	finalsheet = finalbook.active
	finalsheet.append(["Sentence Date", "Disposition Method", "First Name", "Middle Name", "Last Name", "County", "Crime Code", "Severity"])

	for i in progressbar(range(2010, 2022), "Working: ", 40):	# create list of files by year
		files = sorted(glob.glob("xlsx_out/" + str(i) + "-*.xlsx"))

		combineSheets(files, i)

		files = sorted(glob.glob("xlsx_out/combined/" + str(i) + ".xlsx"))
		for file in files:
			parseFile(file, finalsheet)

	finalbook.save(filename="fixed/compiled.xlsx")
	finalbook.close()

if __name__ == "__main__": main()
